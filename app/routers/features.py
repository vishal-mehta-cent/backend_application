# backend/app/routers/features.py

"""
Feature gating allowlist (CSV or Excel).

Supported columns (header row):
  username | email | phone | allow_generate_signals | allow_chart_recommendation | allow_recommendation_page

Where to store the file:
- Render (recommended): /data/feature_access.csv   (or .xlsx)
- Local:               backend/app/data/feature_access.csv (or .xlsx)

You can also force a specific path with env:
  FEATURE_ACCESS_FILE=/data/feature_access.csv
"""

from __future__ import annotations

from fastapi import APIRouter
from typing import Any, Dict, List, Optional
import os
import re
import sqlite3
import csv

from openpyxl import load_workbook

router = APIRouter(prefix="/features", tags=["features"])

# =====================================================
# DB PATH — Render + local (same approach as users.py)
# =====================================================
IS_RENDER = bool(os.getenv("RENDER")) or os.path.isdir("/data")

BACKEND_ROOT = os.path.dirname(
    os.path.dirname(
        os.path.dirname(os.path.abspath(__file__))
    )
)  # .../backend

LOCAL_DB = os.path.join(BACKEND_ROOT, "app", "data", "paper_trading.db")
RENDER_DB = "/data/paper_trading.db"

def _primary_db_path() -> str:
    return RENDER_DB if IS_RENDER else LOCAL_DB

DB_PATH = _primary_db_path()

# =====================================================
# FEATURE ACCESS FILE PATH (CSV or XLSX)
# Priority:
#  1) FEATURE_ACCESS_FILE env (if exists)
#  2) Render disk (/data) CSV then XLSX
#  3) Local backend/app/data CSV then XLSX
# =====================================================
APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # .../backend/app

LOCAL_CSV = os.path.join(APP_DIR, "data", "feature_access.csv")
LOCAL_XLSX = os.path.join(APP_DIR, "data", "feature_access.xlsx")

RENDER_CSV = "/data/feature_access.csv"
RENDER_XLSX = "/data/feature_access.xlsx"

FEATURE_ACCESS_FILE = (os.getenv("FEATURE_ACCESS_FILE") or "").strip()

_CANDIDATES = [
    FEATURE_ACCESS_FILE,
    RENDER_CSV,
    RENDER_XLSX,
    LOCAL_CSV,
    LOCAL_XLSX,
]

def _pick_feature_file() -> str:
    for p in _CANDIDATES:
        if p and os.path.exists(p):
            return p
    # fallback (useful for debug output)
    return FEATURE_ACCESS_FILE or (RENDER_CSV if IS_RENDER else LOCAL_CSV)

FEATURE_PATH = _pick_feature_file()

# =====================================================
# Cache (reload only if file changed OR path changed)
# =====================================================
_CACHE: Dict[str, Any] = {"mtime": None, "rows": [], "path": None}

def _norm(s: Optional[str]) -> str:
    return (s or "").strip().lower().replace(" ", "")

def _digits(s: Optional[str]) -> str:
    return re.sub(r"\D+", "", (s or "").strip())

def _boolish(v: Any) -> bool:
    # Handles: True/False, "TRUE", "False", 1, 0, 1.0, NaN, None
    if v is None:
        return False
    if isinstance(v, bool):
        return bool(v)
    if isinstance(v, (int, float)):
        try:
            return float(v) != 0.0
        except Exception:
            return False
    s = str(v).strip().lower()
    if s in {"", "none", "nan", "null"}:
        return False
    return s in {"1", "true", "t", "yes", "y", "on", "allow", "allowed"}

def _ensure_keys(r: Dict[str, Any]) -> Dict[str, Any]:
    # normalize keys to lowercase and strip spaces
    return {str(k).strip().lower(): v for k, v in (r or {}).items()}

def _read_csv_rows(path: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    with open(path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if not row:
                continue
            r = _ensure_keys(row)
            out.append(
                {
                    "username": _norm(str(r.get("username", ""))),
                    "email": _norm(str(r.get("email", ""))),
                    "phone": _digits(str(r.get("phone", ""))),
                    "allow_generate_signals": _boolish(r.get("allow_generate_signals")),
                    "allow_chart_recommendation": _boolish(r.get("allow_chart_recommendation")),
                    "allow_recommendation_page": _boolish(r.get("allow_recommendation_page")),
                    "_source_file": path,
                }
            )
    return out

def _read_xlsx_rows(path: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [str(h).strip().lower() if h is not None else "" for h in header_cells]
    idx = {name: i for i, name in enumerate(headers) if name}

    def get(row, col, default=None):
        i = idx.get(col)
        if i is None or i >= len(row):
            return default
        return row[i]

    out: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        out.append(
            {
                "username": _norm(str(get(row, "username", ""))),
                "email": _norm(str(get(row, "email", ""))),
                "phone": _digits(str(get(row, "phone", ""))),
                "allow_generate_signals": _boolish(get(row, "allow_generate_signals")),
                "allow_chart_recommendation": _boolish(get(row, "allow_chart_recommendation")),
                "allow_recommendation_page": _boolish(get(row, "allow_recommendation_page")),
                "_source_file": path,
            }
        )
    return out

def _read_feature_rows() -> List[Dict[str, Any]]:
    """
    Reads allowlist from FEATURE_PATH (or best available candidate).
    Caches by (path, mtime).
    """
    global FEATURE_PATH
    FEATURE_PATH = _pick_feature_file()
    path = FEATURE_PATH

    if not path or not os.path.exists(path):
        # no file => no access for anyone
        _CACHE["path"] = path
        _CACHE["mtime"] = None
        _CACHE["rows"] = []
        return []

    try:
        mtime = os.path.getmtime(path)
    except Exception:
        mtime = None

    if (
        _CACHE.get("path") == path
        and _CACHE.get("mtime") == mtime
        and isinstance(_CACHE.get("rows"), list)
    ):
        return _CACHE["rows"]

    # read file fresh
    rows: List[Dict[str, Any]] = []
    try:
        if path.lower().endswith(".csv"):
            rows = _read_csv_rows(path)
        else:
            rows = _read_xlsx_rows(path)
    except Exception:
        rows = []

    _CACHE["path"] = path
    _CACHE["mtime"] = mtime
    _CACHE["rows"] = rows
    return rows

# =====================================================
# Lookup email/phone from users table
# =====================================================
def _lookup_user_contact(username: str) -> Dict[str, str]:
    username = (username or "").strip()
    if not username:
        return {"email": "", "phone": ""}

    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT email, phone FROM users WHERE username=?", (username,))
        row = c.fetchone()
        if not row:
            return {"email": "", "phone": ""}
        return {"email": _norm(row[0]), "phone": _digits(row[1])}
    except Exception:
        return {"email": "", "phone": ""}
    finally:
        try:
            conn.close()
        except Exception:
            pass

# =====================================================
# Public helper used by other routers
# =====================================================
def get_feature_flags(username: str) -> Dict[str, bool]:
    """
    Match order:
    1) username (normalized)
    2) email from users table
    3) phone from users table
    """
    u = _norm(username)
    contact = _lookup_user_contact(username)
    em = contact.get("email", "")
    ph = contact.get("phone", "")

    flags = {
        "allow_generate_signals": False,
        "allow_chart_recommendation": False,
        "allow_recommendation_page": False,
    }

    for r in _read_feature_rows():
        if (
            (u and r.get("username") and r.get("username") == u)
            or (em and r.get("email") and r.get("email") == em)
            or (ph and r.get("phone") and r.get("phone") == ph)
        ):
            flags["allow_generate_signals"] = bool(r.get("allow_generate_signals"))
            flags["allow_chart_recommendation"] = bool(r.get("allow_chart_recommendation"))
            flags["allow_recommendation_page"] = bool(r.get("allow_recommendation_page"))
            break

    return flags

# =====================================================
# APIs
# =====================================================
@router.get("/access/{username}")
def get_access(username: str) -> Dict[str, Any]:
    return {"username": username, **get_feature_flags(username)}

@router.get("/debug/{username}")
def debug_access(username: str) -> Dict[str, Any]:
    """
    Useful to confirm which file is being read on Render/local.
    """
    u = _norm(username)
    rows = _read_feature_rows()
    flags = get_feature_flags(username)

    # show match if any
    contact = _lookup_user_contact(username)
    em = contact.get("email", "")
    ph = contact.get("phone", "")

    matched_row = None
    for r in rows:
        if (
            (u and r.get("username") == u)
            or (em and r.get("email") == em)
            or (ph and r.get("phone") == ph)
        ):
            matched_row = r
            break

    return {
        "username": username,
        "normalized_username": u,
        "is_render": IS_RENDER,
        "db_path": DB_PATH,
        "feature_path_used": FEATURE_PATH,
        "feature_file_exists": bool(FEATURE_PATH and os.path.exists(FEATURE_PATH)),
        "cache_path": _CACHE.get("path"),
        "cache_mtime": _CACHE.get("mtime"),
        "rows_loaded": len(rows),
        "user_email_from_db": em,
        "user_phone_from_db": ph,
        "matched_row": matched_row,
        "flags": flags,
    }
